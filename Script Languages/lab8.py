import argparse
import csv
import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font


def read_csv(file_name):
    with open(file_name, newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        data = [row for row in reader]
    return data


def calculate_stats(data):
    total_price = 0
    total_count = 0
    state_count = defaultdict(int)

    for row in data:
        price = float(row['price'])
        total_price += price
        total_count += 1
        state_count[row['state']] += 1

    avg_price = total_price / total_count
    return avg_price, state_count, total_count


def write_excel(output_file, data, avg_price, state_count, total):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Report'

    header_font = Font(bold=True)

    ws.cell(row=1, column=1, value='State').font = header_font
    ws.cell(row=1, column=2, value='Count').font = header_font

    for i, (state, count) in enumerate(state_count.items(), 1):
        ws.cell(row=2 + i, column=1, value=state)
        ws.cell(row=2 + i, column=2, value=count)

    ws.cell(row=2 + i + len(state_count) + 2, column=1, value='Average Price').font = header_font
    ws.cell(row=2 + i  + len(state_count) + 2, column=2, value=avg_price)

    ws.cell(row=2 + i  + len(state_count) + 3, column=1, value='Total Properties').font = header_font
    ws.cell(row=2 + i  + len(state_count) + 3, column=2, value=total)

    wb.save(output_file)


def main():
    parser = argparse.ArgumentParser(description='Process a real estate dataset.')
    parser.add_argument('dataset', help='The dataset file in CSV format.')
    parser.add_argument('-o', '--output', metavar='OUTPUT', help='The output Excel file.')
    args = parser.parse_args()

    if not args.dataset.endswith('.csv'):
        print('Error: The dataset file must have a .csv extension.')
        return

    if not os.path.exists(args.dataset):
        print('Error: The dataset file does not exist.')
        return

    data = read_csv(args.dataset)
    avg_price, state_count, total = calculate_stats(data)

    if args.output:
        write_excel(args.output, data, avg_price, state_count, total)
    else:
        print(f'Total properties: {total}')


if __name__ == '__main__':
    main()
